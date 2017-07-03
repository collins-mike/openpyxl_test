'''
Created on Jul 2, 2017

@author: Mike
'''
from openpyxl import *
from openpyxl.drawing.image import Image


wb = Workbook()

# grab the active worksheet
ws = wb.active
#create image
img=Image('images/lasercat.jpg')


ws.add_image(img,'C1')
# Data can be assigned directly to cells
for i in range(1,100):
    ws['A'+str(i)] = i

# Python types will automatically be converted
import datetime
ws['B1'] = "=SUM(A1:A100)"

# Save the file
wb.save("sample.xlsx")







if __name__ == '__main__':
    pass